import openpyxl

book = openpyxl.load_workbook("/Users/sankalp/PycharmProjects/PythonSelFramework/PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
firstname = sheet.cell(row=2, column=2).value = "Sankalp"
print(firstname)

print(sheet.max_row)
print(sheet.max_column)

#another way to print values

print(sheet['A3'].value)
print("This is new line")


print("This is a line gitDEMO")
print("This is git - jenkins")